import openpyxl
from collections import defaultdict

column_dict = dict(caja=1, año=2, prod=3, titulo=4,
             autor=5, director=6, escena=7, luz=8,
             vestido=9, notas=10)

problem_columns = ['autor', 'director', 'escena', 'luz', 'vestido']
seen_names = dict()
name_results = dict()
for item in problem_columns:
    seen_names[item] = []
    name_results[item] = []

def parse_names(string):
    split_string = [name.strip() for name in string.split(',')]
    connecting_words = ['del', 'de', 'la', 'y', '-', 'di']

    names = []

    if split_string == ['']: split_string = []

    for name in split_string:
        result = dict()
        name_split = name.split()

        result['primero'] = name_split[0]

        if len(name_split) == 1:
            result['apellido'] = 'N/A'

        elif len(name_split) == 2:
            result['apellido'] = name_split[1]

        else:
            if name_split[1] in connecting_words:
                result['apellido'] = ' '.join(name_split[1:])

            else:
                result['segundo'] = name_split[1]
                result['apellido'] = ' '.join(name_split[2:])

        names.append(result)

    return names

def insert_people(key, entity_table, insert_table, entity_id):
    template_people = 'INSERT INTO {} (primer_nombre, segundo_nombre, apellido) VALUES {};\n\n'
    template_insert = 'INSERT INTO {} ({}_id, obra_id) VALUES ({}, {});\n\n'

    p_id = 0
    with open('{}.sql'.format(key), 'w') as f:
        for person in name_results[key]:
            p_id += 1
            for val in person.values():
                c = val['parse']
                primer_nombre = c.get('primero', '')
                segundo_nombre = c.get('segundo', '')
                apellido = c.get('apellido', '')
                obras = (val['obras'])

            f.write(template_people.format(entity_table, (primer_nombre, segundo_nombre, apellido)))
            for obra in obras:
                f.write(template_insert.format(insert_table, entity_id, p_id, obra))


def table_parse(f):
    results = dict()
    for key in column_dict.keys():
        results[key] = []

    wb = openpyxl.load_workbook(f)
    sh = wb.get_sheet_by_name('Sheet4')

    for i in range(1, sh.max_row + 1):
        for key in results.keys():
            v = sh.cell(row=i, column=column_dict[key]).value

            if key not in problem_columns:
                if v is not None and type(v) != int and v.isdigit():
                    v = int(v)

                results[key].append(v if v is not None else '')

            else:
                v = parse_names(v) if v is not None else ''
                if v != '':
                    for name in v:
                        full_name = ' '.join(list(name.values()))
                        if full_name not in seen_names[key]:
                            seen_names[key].append(full_name)
                            if f == 'universitario.xlsx':
                                name_results[key].append({full_name: {'parse':name, 'obras':[i + 33]}})

                            elif f == 'rodante.xlsx':
                                name_results[key].append({full_name: {'parse':name, 'obras':[i + 33 + 305]}})

                            else:
                                name_results[key].append({full_name: {'parse':name, 'obras':[i]}})

                        else:
                            for item in name_results[key]:
                                if full_name in item.keys():
                                    if f == 'universitario.xlsx':
                                        item[full_name]['obras'].append(i + 33)

                                    elif f == 'rodante.xlsx':
                                        item[full_name]['obras'].append(i + 33 + 305)

                                    else:
                                        item[full_name]['obras'].append(i)
    return results


def make_database(t1, t2, t3):
    template_obras = 'INSERT INTO {} (teatro_id, caja, num_de_prod, titulo, año, notas) VALUES {};\n\n'

    caja1 = t1['caja']
    prod1 = t1['prod']
    titulo1 = t1['titulo']
    year1 = t1['año']
    notas1 = t1['notas']
    teatro1 = [1] * len(caja1)
    directors1 = t1['director']
    autores1 = t1['autor']

    caja2 = t2['caja']
    prod2 = t2['prod']
    titulo2 = t2['titulo']
    year2 = t2['año']
    notas2 = t2['notas']
    teatro2 = [2] * len(caja2)
    directors2 = t2['director']
    autores2 = t2['autor']


    caja3 = t3['caja']
    prod3 = t3['prod']
    titulo3 = t3['titulo']
    year3 = t3['año']
    notas3 = t3['notas']
    teatro3 = [3] * len(caja3)
    directors3 = t3['director']
    autores3 = t3['autor']

    # Insert into Obras
    payload1 = zip(teatro1, caja1, prod1, titulo1, year1, notas1)
    payload2 = zip(teatro2, caja2, prod2, titulo2, year2, notas2)
    payload3 = zip(teatro3, caja3, prod3, titulo3, year3, notas3)

    with open('query.sql', 'w') as f:
        f.write('INSERT INTO Teatros (teatro_id, nombre_teatro) VALUES (1, \'Teatro Comedieta\');\n\n')
        f.write('INSERT INTO Teatros (teatro_id, nombre_teatro) VALUES (2, \'Teatro Universitario\');\n\n')
        f.write('INSERT INTO Teatros (teatro_id, nombre_teatro) VALUES (3, \'Teatro Rodante\');\n\n')

        for query in list(payload1):
            f.write(template_obras.format('Obras', query))

        for query in list(payload2):
            f.write(template_obras.format('Obras', query))

        for query in list(payload3):
            f.write(template_obras.format('Obras', query))

        insert_people('director', 'Directores', 'Directores_de_Obras', 'director')
        insert_people('autor', 'Autores', 'Autores_de_Obras', 'autor')
        insert_people('escena', 'Diseñadores_de_Escenografia', 'Diseño_de_Escenografia', 'd')
        insert_people('luz', 'Diseñadores_de_Luces', 'Diseño_de_Luces', 'd')
        insert_people('vestido', 'Diseñadores_de_Vestuario', 'Diseño_de_Vestuario', 'd')


if __name__ == '__main__':
    comedieta = table_parse('comedieta.xlsx')
    universitario = table_parse('universitario.xlsx')
    rodante = table_parse('rodante.xlsx')

    make_database(comedieta, universitario, rodante)

    print("Parse complete\n")
