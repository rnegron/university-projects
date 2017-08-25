# Raúl Negrón
# -*- coding: utf-8 -*-

import openpyxl

# Dictionary for the Obras table columns
column_dict = dict(caja=1, año=2, prod=3, titulo=4,
                   autor=5, director=6, escena=7, luz=8,
                   vestido=9, notas=10)

# List containing column aliases where the data consists of names
problem_columns = ['autor', 'director', 'escena', 'luz', 'vestido']

# List containing common naming symbols that are still part of a name
connecting_words = ['del', 'de', 'la', 'y', '-', 'di']

# Dictionaries for keeping track of names
seen_names = dict()
name_results = dict()

# Populate both dictionaries
for item in problem_columns:
    seen_names[item] = []
    name_results[item] = []


def parse_names(string):
    '''
    Input: an unformated string containing a name record
    Output: a list with parsed names as elements

    Parsing consists of splitting first names, middle names and
    last names as well as checking for empty collumns
    '''

    # Seperate the unparsed string by comma
    split_string = [name.strip() for name in string.split(',')]

    names = []

    # Fix for empty columns
    if split_string == ['']:
        split_string = []

    # Loop for each individual person
    for name in split_string:
        result = dict()
        # Split the person's name on the whitespace
        name_split = name.split()

        # The first name will always be the first element
        result['primero'] = name_split[0]

        # If there is only one word, there is no last name
        if len(name_split) == 1:
            result['apellido'] = 'N/A'

        # If there are exactly two words, the last one is the last name
        elif len(name_split) == 2:
            result['apellido'] = name_split[1]

        # There are 3 or more words, parse the name
        else:
            # If the second word is a connecting word, it is still part of the
            # last name
            if name_split[1] in connecting_words:
                result['apellido'] = ' '.join(name_split[1:])

            # If not, we will consider the second word to be the middle name,
            # and the rest is the last name
            else:
                result['segundo'] = name_split[1]
                result['apellido'] = ' '.join(name_split[2:])

        names.append(result)

    return names


def insert_people(f, key, entity_table, insert_table, entity_id):
    '''
    Input: file pointer and 4 strings unique to each table which will be injected
            into the templates
    Output: None
    '''

    # Create MySQL template for inserting into a generic table for names
    template_people = 'INSERT INTO {} (primer_nombre, segundo_nombre, apellido) VALUES {};\n\n'

    # Create MySQL template for inserting into a table which links names to
    # Obras
    template_insert = 'INSERT INTO {} ({}_id, obra_id) VALUES ({}, {});\n\n'

    # Keep track for assigning unique ID
    p_id = 0

    # There may be multiple names per column, loop
    for person in name_results[key]:
        p_id += 1

        # Parse the names by column
        for val in person.values():
            c = val['parse']
            primer_nombre = c.get('primero', '')
            segundo_nombre = c.get('segundo', '')
            apellido = c.get('apellido', '')

            # Get their related ID for each Obras element
            obras = (val['obras'])

        # Write the name data into the file according to the tempalte
        f.write(
            template_people.format(
                entity_table,
                (primer_nombre,segundo_nombre,apellido)))

        # Write the data corresponding to the connection of the name and Obras
        for obra in obras:
            f.write(
                template_insert.format(
                    insert_table,
                    entity_id,
                    p_id,
                    obra))


def table_parse(f):
    '''
    Input: a string containing the filename to work with
    Output: a dictionary where key=column and value=data
    '''

    # Create and populate the results dict
    results = dict()
    for key in column_dict.keys():
        results[key] = []

    # Open and navigate to the correct spreadsheet
    wb = openpyxl.load_workbook(f)
    sh = wb.get_sheet_by_name('Sheet4')

    # Loop through the rows
    for i in range(1, sh.max_row + 1):
        # Loop through each column
        for key in results.keys():

            # Get the data in the specific (row, column)
            v = sh.cell(row=i, column=column_dict[key]).value

            # If not a problem (not a data related to names)
            if key not in problem_columns:

                # Verify data integrity
                if v is not None and not isinstance(v, int) and v.isdigit():
                    v = int(v)

                # Append data the results
                results[key].append(v if v is not None else '')

            # If it is problem (it is a data related to names)
            else:

                # Get the parsed data in the specific (row, column)
                v = parse_names(v) if v is not None else ''

                # Only work on it if not empty
                if v != '':
                    # Loop through every word found
                    for name in v:

                        # Create the full name from the separate column
                        full_name = ' '.join(list(name.values()))

                        # If the name is new, add it to results
                        if full_name not in seen_names[key]:
                            seen_names[key].append(full_name)

                            # Make sure the unique identifier for Obras is
                            # correct for each table

                            if f == 'universitario.xlsx':
                                name_results[key].append(
                                    {full_name: {'parse': name, 'obras': [i + 33]}})

                            elif f == 'rodante.xlsx':
                                name_results[key].append(
                                    {full_name: {'parse': name, 'obras': [i + 33 + 305]}})

                            else:
                                name_results[key].append(
                                    {full_name: {'parse': name, 'obras': [i]}})

                        # If the name is not new, find the cache and add to it
                        else:
                            for item in name_results[key]:

                                # Found the name, add the new values to it
                                if full_name in item.keys():

                                    # Make sure the unique identifier for Obras
                                    # is correct for each table
                                    if f == 'universitario.xlsx':
                                        item[full_name]['obras'].append(i + 33)

                                    elif f == 'rodante.xlsx':
                                        item[full_name]['obras'].append(
                                            i + 33 + 305)

                                    else:
                                        item[full_name]['obras'].append(i)
    return results


def make_database(t1, t2, t3):
    '''
    Input: Parsed data from 3 tables
    Output: None

    Sequentially writes the parsed data to the resulting
    MySQL file for database creation.
    '''

    # Template for inserting all the elements of Obras
    template_obras = 'INSERT INTO {} (teatro_id, caja, num_de_prod, titulo, año, notas) VALUES {};\n\n'

    # Get the parsed data from the first table
    caja1 = t1['caja']
    prod1 = t1['prod']
    titulo1 = t1['titulo']
    year1 = t1['año']
    notas1 = t1['notas']
    teatro1 = [1] * len(caja1)
    directors1 = t1['director']
    autores1 = t1['autor']

    # Get the parsed data from the second table
    caja2 = t2['caja']
    prod2 = t2['prod']
    titulo2 = t2['titulo']
    year2 = t2['año']
    notas2 = t2['notas']
    teatro2 = [2] * len(caja2)
    directors2 = t2['director']
    autores2 = t2['autor']

    # Get the parsed data from the third table
    caja3 = t3['caja']
    prod3 = t3['prod']
    titulo3 = t3['titulo']
    year3 = t3['año']
    notas3 = t3['notas']
    teatro3 = [3] * len(caja3)
    directors3 = t3['director']
    autores3 = t3['autor']

    # Bundle the obtained data for each spreadsheet table into payloads
    payload1 = zip(teatro1, caja1, prod1, titulo1, year1, notas1)
    payload2 = zip(teatro2, caja2, prod2, titulo2, year2, notas2)
    payload3 = zip(teatro3, caja3, prod3, titulo3, year3, notas3)

    # Open a results file for writing the data
    with open('query.sql', 'w') as f:
        # Insert Teatros first due to Foreign Key constraint
        f.write(
            'INSERT INTO Teatros (teatro_id, nombre_teatro) VALUES (1, \'Teatro Comedieta\');\n\n')
        f.write(
            'INSERT INTO Teatros (teatro_id, nombre_teatro) VALUES (2, \'Teatro Universitario\');\n\n')
        f.write(
            'INSERT INTO Teatros (teatro_id, nombre_teatro) VALUES (3, \'Teatro Rodante\');\n\n')

        # Insert every element of Obras for all 3 spreadsheets
        for query in list(payload1):
            f.write(template_obras.format('Obras', query))

        for query in list(payload2):
            f.write(template_obras.format('Obras', query))

        for query in list(payload3):
            f.write(template_obras.format('Obras', query))

        # Finally, insert all names for each specific role
        insert_people(
            f,
            'director',
            'Directores',
            'Directores_de_Obras',
            'director')
        insert_people(f, 'autor', 'Autores', 'Autores_de_Obras', 'autor')
        insert_people(
            f,
            'escena',
            'Diseñadores_de_Escenografia',
            'Diseño_de_Escenografia',
            'd')
        insert_people(f, 'luz', 'Diseñadores_de_Luces', 'Diseño_de_Luces', 'd')
        insert_people(
            f,
            'vestido',
            'Diseñadores_de_Vestuario',
            'Diseño_de_Vestuario',
            'd')


if __name__ == '__main__':
    # Retrieve and parse preliminary data from the spreadsheets
    comedieta = table_parse('comedieta.xlsx')
    universitario = table_parse('universitario.xlsx')
    rodante = table_parse('rodante.xlsx')

    # Fix the data and write it to a file in MySQL format
    make_database(comedieta, universitario, rodante)

    print("Parse complete\n")
